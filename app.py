import datetime
import logging
import os
import threading
import time
from pathlib import Path

from flask import Flask, jsonify, render_template
from openpyxl import Workbook, load_workbook

from monitor_selenium import (
    POLL_INTERVAL_SECONDS,
    URL,
    close_driver,
    scrape_products,
    setup_driver,
)
from product_tracker import ProductTracker

LOGGER = logging.getLogger(__name__)
LOG_PATH = Path(__file__).with_name("log.xlsx")

app = Flask(__name__)

tracker = ProductTracker()
state_lock = threading.Lock()
last_updated = "Never"
last_error = None
is_scraping = False


def log_to_excel(name, link, status):
    try:
        if LOG_PATH.exists():
            workbook = load_workbook(LOG_PATH)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Date", "Time", "Status", "Product Name", "Link"])

        now = datetime.datetime.now()
        worksheet.append(
            [
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                status,
                name,
                link,
            ]
        )
        workbook.save(LOG_PATH)
        LOGGER.info("Logged %s alert for %s", status, name)
    except Exception:
        LOGGER.exception("Failed to append alert to %s", LOG_PATH)


def scraper_loop():
    global is_scraping, last_error, last_updated

    LOGGER.info("Starting background scraper for %s", URL)
    driver = None

    try:
        while True:
            started_at = time.monotonic()
            with state_lock:
                is_scraping = True

            try:
                if driver is None:
                    LOGGER.info("Initializing WebDriver")
                    driver = setup_driver()

                products, _ = scrape_products(driver)
                now = datetime.datetime.now()
                with state_lock:
                    events = tracker.update(products, now=now)
                    last_updated = now.strftime("%Y-%m-%d %H:%M:%S")
                    last_error = None

                for event in events:
                    status = "NEW_STOCK" if event["type"] == "NEW" else "RESTOCK"
                    product = event["product"]
                    log_to_excel(product["name"], product["link"], status)
            except Exception as exc:
                LOGGER.exception("Scrape failed; retaining the previous snapshot")
                with state_lock:
                    last_error = str(exc)
                close_driver(driver)
                driver = None
            finally:
                with state_lock:
                    is_scraping = False

            duration = time.monotonic() - started_at
            sleep_seconds = max(1.0, POLL_INTERVAL_SECONDS - duration)
            LOGGER.info(
                "Scrape cycle finished in %.1fs; sleeping %.1fs",
                duration,
                sleep_seconds,
            )
            time.sleep(sleep_seconds)
    finally:
        close_driver(driver)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/data")
def get_data():
    with state_lock:
        snapshot = tracker.snapshot()
        snapshot.update(
            {
                "last_updated": last_updated,
                "last_error": last_error,
                "is_scraping": is_scraping,
            }
        )
    return jsonify(snapshot)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    thread = threading.Thread(target=scraper_loop, daemon=True)
    thread.start()

    debug = os.getenv("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(debug=debug, use_reloader=False)
